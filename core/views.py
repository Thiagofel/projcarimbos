from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.urls import reverse
from django.contrib import messages
from .models import Cliente, Carimbo
from django.core.exceptions import ValidationError
from django.conf import settings
import os
import pandas as pd
from fuzzywuzzy import process, fuzz
from openpyxl import load_workbook
from .forms import ClienteForm

def index(request):
    form = ClienteForm()

    context = {
        'form' : form
    }
    return render(request, 'index.html', context)

def carimbo(request):
    form = ClienteForm(request.POST or None)
    
    if request.method == 'POST' and form.is_valid():
        nome = form.cleaned_data['nome']
        data_nascimento = form.cleaned_data['data_nascimento']

        # Verifique se existe um cliente com o nome exato e data de nascimento
        cliente = Cliente.objects.filter(nome=nome, data_nascimento=data_nascimento).first()

        # Se não encontrar, sugira nomes similares
        if not cliente:
        # Busque todos os nomes no banco
            todos_clientes = Cliente.objects.all()
            nomes_cadastrados = [cliente.nome for cliente in todos_clientes]
             # Encontre os nomes com uma similaridade acima de 80%
            sugestoes = process.extract(nome, nomes_cadastrados, limit=4, scorer=fuzz.partial_ratio)
            nome_sugerido = [sugestao[0] for sugestao in sugestoes if sugestao[1] >= 85]

        # Envia uma sugestão ao usuário, caso exista
            return render(request, 'index.html', {
                'form' : form,
                'nome_sugerido': nome_sugerido,
                'erro': "Cadastro não encontrado. Seu nome pode ter sido digitado errado. Por favor, tente novamente com as possíveis sugestões abaixo e avise o responsável pelo whatsapp." if nome_sugerido else "Nome não encontrado.",
            })  
        
        # Se o cliente existe, continue com o fluxo de carimbos
        carimbo, _ = Carimbo.objects.get_or_create(cliente=cliente)

        # Determina uma mensagem de notificação
        notificacao = ""
        if carimbo.carimbo == 4:
            notificacao = "Você está prestes a completar a cartela!"
        elif carimbo.carimbo == 5:
            notificacao = "Parabéns! Você completou a cartela."

        # Escolhe a imagem com base na quantidade de carimbos
        imagem_carimbo = f"images/carimbo{carimbo.carimbo}.png"

        # Armazena os dados na sessão para usar no redirecionamento
        request.session['nome'] = cliente.nome
        request.session['carimbo'] = carimbo.carimbo
        request.session['notificacao'] = notificacao
        request.session['imagem_carimbo'] = imagem_carimbo
        request.session['login'] = cliente.login
        request.session['senha_login'] = cliente.senha_login

        # Redireciona para evitar o reenvio do formulário
        return redirect(reverse('carimbo'))

    # Exibe a página usando os dados armazenados na sessão
    nome = request.session.get('nome')
    carimbo = request.session.get('carimbo')
    notificacao = request.session.get('notificacao')
    imagem_carimbo = request.session.get('imagem_carimbo')
    login = request.session.get('login')
    senha_login = request.session.get('senha_login')

    # Passa o cliente, total de carimbos e notificação para o template
    return render(request, 'carimbo.html', {
        'nome': nome,
        'carimbo': carimbo,
        'notificacao': notificacao,
        'imagem_carimbo': imagem_carimbo,
        'login': login,
        'senha_login': senha_login
    })
    
def importar_dados(request):
    try:
        # Defina o caminho absoluto para o arquivo Excel
        file_path = os.path.join(settings.BASE_DIR,'projnath.xlsx')

        # Verifique se o arquivo existe
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")

        # Leia o arquivo Excel usando pandas
        df = pd.read_excel(file_path)

        # Verifique se as colunas necessárias estão presentes no arquivo
        colunas_requeridas = ['nome', 'data_nascimento', 'login', 'senha_login']
        if not all(col in df.columns for col in colunas_requeridas):
            raise ValidationError("O arquivo Excel não contém todas as colunas necessárias.")

        # Converta 'data_nascimento' para datetime, caso não esteja no formato correto
        df['data_nascimento'] = pd.to_datetime(df['data_nascimento'], errors='coerce')

        # Verifique se há valores de data inválidos após a conversão
        if df['data_nascimento'].isnull().any():
            raise ValidationError("Algumas datas de nascimento estão no formato inválido.")
        
        # Itera sobre as linhas e salva ou atualiza os dados no banco
        for _, row in df.iterrows():
            # Certifique-se de que a data está no formato correto antes de salvar
            Cliente,  = Cliente.objects.update_or_create(
                nome=row['nome'],
                data_nascimento=row['data_nascimento'].date(),
                defaults={
                    'login': row['login'],
                    'senha_login': row['senha_login']
                }
            )

            if Cliente:
                messages.info(request, f"Cliente {row['nome']} atualizado com sucesso.")                   

        messages.success(request, 'Dados importados e atualizados com sucesso!')
        return redirect('index')  # Redireciona após o sucesso

    except FileNotFoundError as e:
        messages.error(request, f"Erro: {str(e)}")
    except ValidationError as e:
        messages.error(request, f"Erro: {str(e)}")
    except Exception as e:
        messages.error(request, f"Erro ao importar os dados: {str(e)}")

    return HttpResponse(status=200)  # Retorna um status de sucesso sem renderizar template

# def importar_dados(request):
    try:
        # Defina o caminho absoluto para o arquivo Excel
        file_path = os.path.join(settings.BASE_DIR, 'media', 'exce_files', 'projnath.xlsx')

        # Verifique se o arquivo existe
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")

        # Carrega o arquivo Excel
        workbook = load_workbook(file_path)
        sheet = workbook.active  # Usa a planilha ativa (primeira)

        # Verifique se as colunas necessárias estão presentes
        colunas_requeridas = ['nome', 'data_nascimento', 'login', 'senha_login']
        headers = [cell.value for cell in sheet[1]]  # Primeira linha como cabeçalhos
        if not all(col in headers for col in colunas_requeridas):
            raise ValidationError("O arquivo Excel não contém todas as colunas necessárias.")

        # Mapeia os índices das colunas
        col_indices = {col: headers.index(col) for col in colunas_requeridas}

        # Itera sobre as linhas (a partir da segunda linha)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            nome = row[col_indices['nome']]
            data_nascimento = row[col_indices['data_nascimento']]
            login = row[col_indices['login']]
            senha_login = row[col_indices['senha_login']]

            # Verifica se a data é válida
            if not isinstance(data_nascimento, (str, type(None))):
                data_nascimento = data_nascimento.date()

            # Atualiza ou cria o cliente
            Cliente, = Cliente.objects.update_or_create(
                nome=nome,
                data_nascimento=data_nascimento,
                defaults={
                    'login': login,
                    'senha_login': senha_login
                }
            )

            if Cliente:
                messages.info(request, f"Cliente {nome} atualizado com sucesso.")

        messages.success(request, "Dados importados e atualizados com sucesso!")
        return redirect('index')  # Redireciona após o sucesso

    except FileNotFoundError as e:
        messages.error(request, f"Erro: {str(e)}")
    except ValidationError as e:
        messages.error(request, f"Erro: {str(e)}")
    except Exception as e:
        messages.error(request, f"Erro ao importar os dados: {str(e)}")

    return HttpResponse(status=200)  # Retorna um status de sucesso sem renderizar template